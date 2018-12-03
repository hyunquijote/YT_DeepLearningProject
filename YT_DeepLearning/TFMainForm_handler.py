import sys, os
import ctypes
form_path = os.path.dirname( os.path.abspath( __file__ ) )
module_path = os.path.normpath( form_path+"\..\YT_Prediction")
sys.path.insert(0, module_path)
from PyQt5.QtCore       import *
from PyQt5              import uic, QtCore
from PyQt5.QtWidgets    import *
from RcvSckt_handler    import *
from SndSckt_handler    import *
from TensorFlow_handler import *
from CodeDef            import *
from DB_handler         import *
from openpyxl           import load_workbook, Workbook


ui_path = os.path.dirname(os.path.abspath(__file__))
TFMainFormClass = uic.loadUiType(os.path.join(ui_path, "TFMainForm.ui"))[0]

class TFMainForm_handler(QDialog, TFMainFormClass):

    def __init__(self, isUnite="None"):
        super().__init__()

        # 폼 설정
        self.setupUi(self)

        # 시세 변수 초기화
        self.StkList    = []  # 처리종목 리스트
        self.CrPrcDict  = {}  # 현재가/종가 딕셔너리 {(종목코드,시각): 현재가/종가 }
        self.FtPrcDict  = {}  # 시가 딕셔너리        {(종목코드,시각): 현재가/종가 }
        self.HgPrcDict  = {}  # 고가 딕셔너리        {(종목코드,시각): 현재가/종가 }
        self.LoPrcDict  = {}  # 저가 딕셔너리        {(종목코드,시각): 현재가/종가 }

        self.CurMn = datetime.today().strftime("%H%M")  # 처리현재시각(HHMM)

        # 변수 초기화
        self.SndSckt   = None  # 송신소켓 핸들러
        self.RcvSckt   = None  # 수신소켓 핸들러
        self.isRcv     = False # 수신상태
        self.DBH       = None  # DB 핸들러
        self.TFH       = None  # 텐서플로우 핸들러

        # 버튼이벤트 핸들러 설정
        self.btnRcvData.clicked.connect(self.ClickRcvData)               # 데이터수신 버튼
        #self.btnTrmsData.clicked.connect(self.ClickTrmsData)            # 데이터송신 버튼
        self.btnDoLearning.clicked.connect(self.ClickDoLearning)         # 학습수행 버튼
        self.btnDoLearningTest.clicked.connect(self.ClickDoLearningTest) # 학습수행결과테스트 버튼

        # 모델 콤보박스 초기화
        self.cmbMdlTp.addItems(CodeDef.TF_MDL_LIST_RNN + CodeDef.TF_MDL_LIST_STD)

        # 소켓 초기화
        self.SndSckt = SndSckt_handler()
        self.RcvSckt = RcvSckt_handler(self,CodeDef.PORT_TF_DATA)

        # 상태표시
        self.edtStat.setText("시작전")

        # 필요폴더 확인 및 생성, 로그삭제
        self.InitFolder()

        # DB 핸들러 초기화
        self.DBH = DB_handler()

        # TensorFlow 핸들러 초기화
        self.TFH = TensorFlow_handler(self,self.DBH)

        # 학습기간입력 및 TEST일자 초기화
        self.dteStrDtMn.setCalendarPopup(True)
        self.dteStrDtMn.setDisplayFormat("yyyy-MM-dd hh:mm")
        StrDtMn = QtCore.QDateTime( int(CodeDef.TF_LEARNING_STR_DT[ :4]) # 년
                                   ,int(CodeDef.TF_LEARNING_STR_DT[4:6]) # 월
                                   ,int(CodeDef.TF_LEARNING_STR_DT[6: ]) # 일
                                   ,int(CodeDef.TF_LEARNING_STR_MN[ :2]) # 시
                                   ,int(CodeDef.TF_LEARNING_STR_MN[2: ]) # 분
                                   ,0)
        self.dteStrDtMn.setDateTime(StrDtMn)

        self.dteEndDtMn.setCalendarPopup(True)
        self.dteEndDtMn.setDisplayFormat("yyyy-MM-dd hh:mm")
        EndDtMn = QtCore.QDateTime( int(CodeDef.TF_LEARNING_END_DT[ :4])  # 년
                                   ,int(CodeDef.TF_LEARNING_END_DT[4:6])  # 월
                                   ,int(CodeDef.TF_LEARNING_END_DT[6: ])  # 일
                                   ,int(CodeDef.TF_LEARNING_END_MN[ :2])  # 시
                                   ,int(CodeDef.TF_LEARNING_END_MN[2: ])  # 분
                                   ,0)
        self.dteEndDtMn.setDateTime(EndDtMn)

        self.dteTestStrDtMn.setCalendarPopup(True)
        self.dteTestStrDtMn.setDisplayFormat("yyyy-MM-dd hh:mm")
        TestStrDtMn = QtCore.QDateTime(int(CodeDef.TF_TEST_STR_DT[ :4])  # 년
                                     , int(CodeDef.TF_TEST_STR_DT[4:6])  # 월
                                     , int(CodeDef.TF_TEST_STR_DT[6: ])  # 일
                                     , int(CodeDef.TF_TEST_STR_MN[ :2])  # 시
                                     , int(CodeDef.TF_TEST_STR_MN[2: ])  # 분
                                     , 0)
        self.dteTestStrDtMn.setDateTime(TestStrDtMn)

        self.dteTestEndDtMn.setCalendarPopup(True)
        self.dteTestEndDtMn.setDisplayFormat("yyyy-MM-dd hh:mm")
        TestEndDtMn = QtCore.QDateTime(int(CodeDef.TF_TEST_END_DT[ :4])  # 년
                                     , int(CodeDef.TF_TEST_END_DT[4:6])  # 월
                                     , int(CodeDef.TF_TEST_END_DT[6: ])  # 일
                                     , int(CodeDef.TF_TEST_END_MN[ :2])  # 시
                                     , int(CodeDef.TF_TEST_END_MN[2: ])  # 분
                                     , 0)
        self.dteTestEndDtMn.setDateTime(TestEndDtMn)


        # 텐서플로 처리 종목 초기화 ( 입력순서순 정렬되어 있음 )
        # KODEX 레버지리 추가
        self.StkList = list(self.DBH.queryProcStkList()["STK_CD"])
        self.StkList.insert(0,CodeDef.INDEX_STK_CD)
        #print("처리종목확인:",self.StkList)

        # 예측분 리스트 정보
        self.PredMnList = CodeDef.TF_LEARNING_MN_LIST  # 예측분 리스트
        self.PredMnCnt  = len(self.PredMnList)

        # 예측분 시각을 가져오기 위한 임시 일자
        self.TempDt = datetime(year=2000, month=1, day=1, hour=1, minute=1)

        # 예측분 계산을 위한 TimeDelta 리스트
        self.PredTimeDeltaList = []
        for idx in range(self.PredMnCnt):
            self.PredTimeDeltaList.append(timedelta(minutes=self.PredMnList[idx]))

        # 실시간 예측결과 엑셀저장 초기화
        NewRlstFile = False
        self.PredSaveFile = None
        try:
            self.PredSaveFile = load_workbook(form_path+CodeDef.TF_PREDICTION_SAVE_FILE)
        except Exception as e:
            print("실시간시세결과저장엑셀오픈에러:", e)
            self.PredSaveFile = Workbook()
            NewRlstFile = True
            print("엑셀파일 재생성완료")

        self.PredWrkSheet = self.PredSaveFile.active
        self.PredCellPos = 1
        self.PredWrkSheet.cell(row=1, column=1, value=datetime.today().strftime("%Y%m%d"))
        if(NewRlstFile):
            self.PredWrkSheet.cell(row=2, column=1, value="Time")
            self.PredWrkSheet.cell(row=2, column=2, value="RealPrice")
            for idx in range(self.PredMnCnt):
                self.PredWrkSheet.cell(row=2, column=(idx+3), value=("Prediction_"+str(self.PredMnList[idx])))
            self.PredCellPos = 3
        else:
            for row in self.PredWrkSheet.rows:
                self.PredCellPos = self.PredCellPos + 1
                if (row[0].value is None):
                    break

        self.PredSaveFile.save(form_path + CodeDef.TF_PREDICTION_SAVE_FILE)

        # 통합실행의 경우 수신준비
        if (isUnite == "TRUE"):
            self.ClickRcvData()

        return None

    def __del__(self):
        # 송신소켓종료
        if (self.SndSckt.GetOnSock() == True):
            self.SndSckt.ClseSckt()
        # 수신소켓종료
        if (self.isRcv):
            print("데이터 수신 쓰레드 중지 시작")
            self.RcvSckt.DoStop()
            self.RcvSckt.join()
        # 실시간예측저장엑셀 종료
        self.PredSaveFile.close()
        return None

    # 데이터수신처리 시작
    def ClickRcvData(self):
        # 중지실행
        if (self.isRcv):
            print("데이터 수신 쓰레드 중지 시작")
            try:
                self.RcvSckt.DoStop()
                self.RcvSckt.join()
            except Exception as e:
                print("데이터 수신 쓰레드 중지 에러: ", e)
                return None
            self.isRcv = False
            self.edtStat.setText("시작전")
            self.btnRcvData.setText("데이터수신시작")
            self.btnRcvData.setStyleSheet("background-color:None")
            print("데이터 수신 쓰레드 종료됨.")
        # 받기시작
        else:

            try:
                # 딥러닝 모델 초기화
                self.TFH.SetMdlSess(self.GetCmbMdlTpValue(), "Y")

                # 수신소켓 시작
                self.RcvSckt.start()
            except Exception as e:
                print("데이터 수신 쓰레드 시작 에러 :", e)
                return None
            self.isRcv = True
            self.edtStat.setText("데이터수신중")
            self.btnRcvData.setText("데이터수신종료")
            self.btnRcvData.setStyleSheet("background-color:rgb(255,020,147)")
            print("데이터 수신 Loop 시작됨.")

        return None


    # 수신 데이터 처리
    # inData   : 수신 테이터 (QTTN|종목코드|시각(HHMM)|현재or종가|시가|고가|저가|E|종목코드....|END)
    def ProcRcvRealQttn(self, inProcTp, inData):

        self.dispText("시세수신:" + inData)

        RcvData = inData.split("|")
        ProcTp  = RcvData[0] # 처리구분
        del RcvData[0]
        DataCnt = len(RcvData)

        print("처리구분 Cnt:",ProcTp, DataCnt)
        # 시세 처리
        if(ProcTp == "QTTN" or ProcTp == "QTTN_TEST"):
            QttnIdx   = 0
            StkCd     = ""
            MN        = ""
            GetTFRslt = False
            # 시세 파싱
            # [종목코드, 시각(HHMM) , 현재or종가 , 시가 , 고가 , 저가 , E , 종목코드.... , END]
            for i in range(DataCnt):

                Data = RcvData.pop(0)

                # 종료확인
                if (Data == "END"):
                    break
                else:
                    # 종목코드
                    if(QttnIdx == CodeDef.REAL_QTTN_COL_STK_CD):
                        StkCd    = Data
                    # 시각(HHMM)
                    elif(QttnIdx == CodeDef.REAL_QTTN_COL_TIME):
                        MN = Data
                        # 시세 시각이 변동될시 예측 시작
                        if(int(MN) > int(self.CurMn) or ProcTp == "QTTN_TEST"):
                            GetTFRslt = True
                    # 현재가/종가
                    elif (QttnIdx == CodeDef.REAL_QTTN_COL_CRPRC):
                        self.CrPrcDict[(StkCd, MN)] = float(Data)
                    # 시가
                    elif (QttnIdx == CodeDef.REAL_QTTN_COL_FTPRC):
                        self.FtPrcDict[(StkCd, MN)] = float(Data)
                    # 고가
                    elif (QttnIdx == CodeDef.REAL_QTTN_COL_HGPRC):
                        self.HgPrcDict[(StkCd, MN)] = float(Data)
                    # 저가
                    elif (QttnIdx == CodeDef.REAL_QTTN_COL_LOPRC):
                        self.LoPrcDict[(StkCd, MN)] = float(Data)

                    # 1개 종목 시세 종료
                    if( Data == "E"):
                    #if (QttnIdx == CodeDef.REAL_QTTN_COL_END):
                        QttnIdx = 0
                    else:
                        QttnIdx += 1

            # ---- 시세파싱 for문 종료 -------
            if(GetTFRslt):
                print("시간이동으로 예측시작! self.CurMn, MN, GetTFRslt:",self.CurMn, MN, GetTFRslt)
                self.DoPrediction(self.CurMn)
                self.CurMn = MN
                print("시간이동으로 예측종료")
        # ---- ProcTp == "QTTN" 종료 -------
        elif(ProcTp == "TEST"):
            print("ProcTp : TEST")
        else:
            print("ProcTp 설정 오류: ",ProcTp)
            return None

        return None

    # 학습수행
    def ClickDoLearning(self):
        try :
            StrDtMn = self.dteStrDtMn.dateTime().toString("yyyyMMddhhmm")
            EndDtMn = self.dteEndDtMn.dateTime().toString("yyyyMMddhhmm")

            if (self.ChkDtMn(StrDtMn, EndDtMn, "학습수행") == False):
                return None

            TestStrDtMn = self.dteTestStrDtMn.dateTime().toString("yyyyMMddhhmm")
            TestEndDtMn = self.dteTestEndDtMn.dateTime().toString("yyyyMMddhhmm")

            if (self.ChkDtMn(TestStrDtMn, TestEndDtMn, "학습결과 Test") == False):
                return None

            # 예측분별 학습
            for idx in range(len(CodeDef.TF_LEARNING_MN_LIST)):
                print(str(CodeDef.TF_LEARNING_MN_LIST[idx])+"분 예측 학습시작")
                self.TFH.DoLearning( self.GetCmbMdlTpValue() , StrDtMn, EndDtMn, idx);

            ctypes.windll.user32.MessageBoxW(0, "학습종료 저장완료", "학습종료", 0)

        except Exception as e:
            print("학습수행 클릭에러:", sys.exc_info()[-1].tb_lineno,"줄", e)

        return None

    # 학습수행결과 테스트
    def ClickDoLearningTest(self):
        try:
            TestStrDtMn = self.dteTestStrDtMn.dateTime().toString("yyyyMMddhhmm")
            TestEndDtMn = self.dteTestEndDtMn.dateTime().toString("yyyyMMddhhmm")

            if(self.ChkDtMn(TestStrDtMn, TestEndDtMn, "학습결과 Test") == False):
                return None

            # 예측분별 학습
            for idx in range(len(CodeDef.TF_LEARNING_MN_LIST)):
                print(str(CodeDef.TF_LEARNING_MN_LIST[idx]) + "분 예측 학습결과 테스트")
                self.TFH.DoLearningTest(self.GetCmbMdlTpValue(), TestStrDtMn, TestEndDtMn, idx);

            ctypes.windll.user32.MessageBoxW(0, "테스트결과 저장완료", "학습결과 테스트 종료", 0)
        except Exception as e:
            print("학습결과 테스트수행 클릭에러:", sys.exc_info()[-1].tb_lineno,"줄", e)

        return None

    # 학습기간입력 오류 확인
    def ChkDtMn(self, inStrDtMn, inEndDtMn, inTpNm):
        if (len(inStrDtMn) != 12 or len(inEndDtMn) != 12):
            ctypes.windll.user32.MessageBoxW(0, inTpNm+" 기간입력형식이 잘못되었습니다.", "입력오류", 0)
            return False

        if (int(inStrDtMn) >= int(inEndDtMn)):
            ctypes.windll.user32.MessageBoxW(0, inTpNm+" 시작일시가 더 큽니다.", "입력오류", 0)
            return False

        return True


    # 예측수행
    # inMn : 처리대상 시각(HHMM)
    def DoPrediction(self, inMn):

        # 예측을 위한 시세 리스트
        Qttn = []

        # 인덱스 현재가
        KospiPrc = 0.0

        StkCnt = len(self.StkList)
        for idx in range(StkCnt):
            StkCd = self.StkList[idx]

            # 입력데이터 설정
            #print("self.CrPrcDict:",self.CrPrcDict)
            if ((StkCd, inMn) in self.CrPrcDict):
                CrPrc = self.CrPrcDict[(StkCd, inMn)]
                Qttn.append(CrPrc)
                # KODEX 레버리지현재가 다음은 시분 시세 맨 앞으로 넣음.
                if(StkCd == CodeDef.INDEX_STK_CD):
                    Qttn.insert(0,float(inMn))
                    KospiPrc = CrPrc
            # 한종목이라도 값이없다면 예측할수없다.
            else:
                print("시세생성전이라 불가 StkCd, idx, inMn: ", StkCd, idx, inMn)
                return None

            #CrPrc = self.CrPrcDict[(StkCd, inMn)]
            #FtPrc = self.FtPrcDict[(StkCd, inMn)]
            #HgPrc = self.HgPrcDict[(StkCd, inMn)]
            #LoPrc = self.LoPrcDict[(StkCd, inMn)]

        print("예측수행!!!: ",inMn, Qttn)
        # 예측치 가져오기
        inPutX = []
        inPutX.append(Qttn)
        for idx in range(self.PredMnCnt):
            # 미래시간계산. 장시간을 벗어나면 Pass
            #NextMn = self.GetNextMn(inMn, idx)
            NextMn = CodeDef.GetNextMn(None, inMn, self.PredMnList[idx])
            if(int(NextMn) > 1519):
                print("예측시간 오버 예측미수행 NextMn:",NextMn)
                P_Value = 0.0
                continue
            P_Value = self.TFH.DoPrediction(self.GetCmbMdlTpValue(),inPutX, idx); # list형 리턴

            print("예측결과!!!["+str(self.PredMnList[idx])+"분]:", NextMn, P_Value)
            if( P_Value == 0.0):
                print("예측값오류 패스 P_Value:", P_Value)
                continue

            #self.dispText("예측결과!!!:" + NextMn +"|" + str(P_Value))

            # 결과값 전송
            if(self.SndSckt.GetOnSock() == False):
                self.SndSckt.CnntSckt(CodeDef.PORT_TF_RCV_RSLT)
            # 예측분|예측시각|값
            SndData = str(self.PredMnList[idx])+"|"+NextMn+"|"+str(P_Value)+"|E|"
            print("SndData:",SndData)
            self.SndSckt.SendData(SndData)

            # 결과값 저장
            self.PredWrkSheet.cell(row=self.PredCellPos, column=1, value=inMn)
            self.PredWrkSheet.cell(row=self.PredCellPos, column=2, value=KospiPrc)
            self.PredWrkSheet.cell(row=(self.PredCellPos+self.PredMnList[idx]), column=(idx+3), value=P_Value)
            self.PredSaveFile.save(form_path+CodeDef.TF_PREDICTION_SAVE_FILE)

        # 결과저장 엑셀 Row 인덱스 증가
        self.PredCellPos = self.PredCellPos + 1

        return P_Value

    # # 다음 시분 가져오기
    # # inStdMn     : 기준시분(HHMM)
    # # inPredMnIdx : 예측분 인덱스
    # def GetNextMn(self,inStdMn, inPredMnIdx):
    #
    #     NextMn = ""
    #
    #     iM = int(inStdMn[:2])
    #     iN = int(inStdMn[2:])
    #
    #     self.TempDt = self.TempDt.replace(hour=iM, minute=iN)
    #
    #     NextMn = (self.TempDt + self.PredTimeDeltaList[inPredMnIdx]).strftime("%H%M")
    #
    #     return NextMn

    # 출력창 출력
    def dispText(self,inText):
        self.edtTest.append(inText)
        return None

    # 학습저장초기화 체크박스 값 가져오기
    def GetInitSave(self):
        return self.chkInitSave.isChecked()

    # 필요 폴더 확인 및 생성
    def InitFolder(self):

        # 학습테스트 결과폴더
        if not os.path.exists(form_path+CodeDef.TF_LEARNING_RSLT_DIR):
            os.mkdir(form_path+CodeDef.TF_LEARNING_RSLT_DIR)

        # 학습로그 폴더
        if not os.path.exists(form_path+CodeDef.TF_LEARNING_LOG_DIR):
            os.mkdir(form_path+CodeDef.TF_LEARNING_LOG_DIR)

        # 학습결과 폴더
        if not os.path.exists(form_path+CodeDef.TF_LEARNING_SAVE_MAINDIR):
            os.mkdir(form_path+CodeDef.TF_LEARNING_SAVE_MAINDIR)

        # 학습결과 폴더 : 예측분별
        # for idx in range(len(CodeDef.TF_LEARNING_MN_LIST)):
        #     Dir = form_path + CodeDef.TF_LEARNING_SAVE_DIR + str(CodeDef.TF_LEARNING_MN_LIST[idx])
        #     if not os.path.exists(Dir):
        #         os.mkdir(Dir)

        # 실시간 예측결과 저장 폴더
        if not os.path.exists(form_path+CodeDef.TF_PREDICTION_SAVE_DIR):
            os.mkdir(form_path+CodeDef.TF_PREDICTION_SAVE_DIR)

        # 로그파일삭제
        FilePath = form_path+ "\TrainingLog"
        FileNmlist = None
        for dirN, subDirN, fileN in os.walk(FilePath):
            FileNmlist = fileN
        for idx in range(len(FileNmlist)):
            os.remove(FilePath + "\\" + FileNmlist[idx])

        return None

    # 모델 콤포박스 값 가져오기
    def GetCmbMdlTpValue(self):
        return str(self.cmbMdlTp.currentText())
